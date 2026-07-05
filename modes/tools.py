import csv
import json
import msvcrt
import re
import time
from pathlib import Path
from shutil import copy2

import requests

from utils.backup import backup_file
from utils.constants import (
    ALIASES_FILE,
    BACKUP_DIR,
    CACHE_FILE,
    EXPORT_DIR,
    RESUME_FILE,
    RETRY_FILE,
    SETTINGS_FILE,
)
from utils.file_utils import data_file, load_json, save_json

from utils.ui import ask, error, success, warning, show_header, show_menu



SETTING_LABELS = {
    "enable_anilist": "Enable AniList Sync",
    "enable_mal": "Enable MyAnimeList Sync",
    "resume_import": "Resume Imports",
    "retry_failed": "Retry Failed Anime",
    "auto_learn_aliases": "Auto Learn Aliases",
    "franchise_sync": "Franchise Sync",
    "use_search_cache": "Use Search Cache",
    "fuzzy_matching": "Fuzzy Matching",
    "interactive_search": "Interactive Search",
    "auto_backup": "Automatic Backup",
    "confirm_before_sync": "Confirm Before Sync",

    # Non-menu/common settings (kept for backwards compatibility)
    "interactive_sync": "Interactive Search",
    "debug": "Debug Mode",
    "search_threshold": "Search Similarity Threshold",
    "search_results": "Maximum Search Results",
    "max_retries": "Maximum Retries",
    "stop_after": "Stop After",
    "stop_after_existing": "Stop After Existing",
    "anilist_per_page": "AniList Page Size",
    "default_status": "AniList Default Status",
    "mal_default_status": "MAL Default Status",
}


# Settings menu definition (grouped) to avoid dumping every key with enumerate/print.
# Format: [(group_title, [(setting_key, setting_label), ...]), ...]
SETTINGS_MENU = [

    (
        "Synchronization",
        [
            ("enable_anilist", "Enable AniList Sync"),
            ("enable_mal", "Enable MyAnimeList Sync"),
            ("resume_import", "Resume Imports"),
            ("retry_failed", "Retry Failed Anime"),
            ("auto_learn_aliases", "Auto Learn Aliases"),
            ("franchise_sync", "Franchise Sync"),
        ],
    ),

    (
        "Search",
        [
            ("use_cache", "Use Search Cache"),
            ("fuzzy_matching", "Fuzzy Matching"),
            ("interactive_search", "Interactive Search"),
        ],
    ),

    (
        "Backup",
        [
            ("auto_backup", "Automatic Backup"),
        ],
    ),

    (
        "User Interface",
        [
            ("confirm_before_sync", "Confirm Before Sync"),
        ],
    ),

    (
        "Advanced",
        [
            ("debug", "Debug Mode"),
            ("search_threshold", "Search Threshold"),
            ("search_results", "Maximum Search Results"),
            ("max_retries", "Maximum Retries"),
            ("anilist_per_page", "AniList Page Size"),
            ("stop_after", "Stop After"),
            ("stop_after_existing", "Stop After Existing"),
            ("default_status", "AniList Default Status"),
            ("mal_default_status", "MAL Default Status"),
        ],
    ),
]



# BASIC_SETTINGS must be right below SETTINGS_MENU.
BASIC_SETTINGS = {
    "enable_anilist",
    "enable_mal",
    "resume_import",
    "retry_failed",
    "auto_learn_aliases",
    "franchise_sync",
    "use_cache",
    "fuzzy_matching",
    "interactive_search",
    "auto_backup",
    "confirm_before_sync",
}

ADVANCED_SETTINGS = {
    "debug",
    "search_threshold",
    "search_results",
    "max_retries",
    "anilist_per_page",
    "stop_after",
    "stop_after_existing",
    "default_status",
    "mal_default_status",
}



# Map flat setting keys to their top-level section in data/settings.json
SETTINGS_KEY_TO_SECTION = {
    # sync
    "enable_anilist": "sync",
    "enable_mal": "sync",
    "resume_import": "sync",
    "retry_failed": "sync",
    "auto_learn_aliases": "sync",
    "franchise_sync": "sync",

    # search
    "use_search_cache": "search",
    "fuzzy_matching": "search",
    "interactive_search": "search",

    # backup
    "auto_backup": "backup",

    # ui
    "confirm_before_sync": "ui",
}


EXPORT_PATH = Path(EXPORT_DIR)

DATA_FILES = [
    ALIASES_FILE,
    CACHE_FILE,
    RETRY_FILE,
    RESUME_FILE,
    SETTINGS_FILE,
    "missing_anilist.json",
]


def ensure_exports():
    EXPORT_PATH.mkdir(parents=True, exist_ok=True)


def export_path(filename):
    ensure_exports()
    return EXPORT_PATH / filename


def export_json(filename, data):
    path = export_path(_with_suffix(filename, "json"))

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    return path


def export_txt(filename, data):
    path = export_path(_with_suffix(filename, "txt"))

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(_txt_lines(data)))
        f.write("\n")

    return path


def export_csv(filename, rows, headers):
    path = export_path(_with_suffix(filename, "csv"))

    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    return path


def export_markdown(filename, rows, headers):
    path = export_path(_with_suffix(filename, "md"))

    with open(path, "w", encoding="utf-8") as f:
        f.write("| " + " | ".join(headers) + " |\n")
        f.write("| " + " | ".join(["---"] * len(headers)) + " |\n")

        for row in rows:
            values = [
                str(row.get(header, "")).replace("\n", " ")
                for header in headers
            ]
            f.write("| " + " | ".join(values) + " |\n")

    return path


def export_html(filename, rows, headers):
    path = export_path(_with_suffix(filename, "html"))

    with open(path, "w", encoding="utf-8") as f:
        f.write("<!DOCTYPE html>\n<html>\n<head>\n")
        f.write("<meta charset=\"utf-8\">\n")
        f.write("<style>\n")
        f.write("body { font-family: -apple-system, sans-serif; margin: 20px; }\n")
        f.write("table { border-collapse: collapse; width: 100%; }\n")
        f.write("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }\n")
        f.write("th { background: #4a90d9; color: white; }\n")
        f.write("tr:nth-child(even) { background: #f5f5f5; }\n")
        f.write("</style>\n</head>\n<body>\n")
        f.write(f"<h2>{filename}</h2>\n")
        f.write("<table>\n<thead>\n<tr>")
        for h in headers:
            f.write(f"<th>{h}</th>")
        f.write("</tr>\n</thead>\n<tbody>\n")
        for row in rows:
            f.write("<tr>")
            for h in headers:
                val = str(row.get(h, "")).replace("\n", "<br>")
                f.write(f"<td>{val}</td>")
            f.write("</tr>\n")
        f.write("</tbody>\n</table>\n</body>\n</html>\n")

    return path


def export_xlsx(filename, rows, headers):
    path = export_path(_with_suffix(filename, "xlsx"))

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        warning("openpyxl not installed. Install it with: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename

    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, len(h) + 4)

    wb.save(path)
    return path


def choose_format():
    xlsx_available = True
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        xlsx_available = False

    while True:
        options = [
            "JSON",
            "CSV",
            "TXT",
            "Markdown",
            "HTML",
            "Excel (.xlsx)" + ("" if xlsx_available else " [dim](requires: pip install openpyxl)[/]"),
            "Cancel",
        ]
        choice = show_menu("Export Format", options)

        if choice == "1":
            return "json"
        if choice == "2":
            return "csv"
        if choice == "3":
            return "txt"
        if choice == "4":
            return "md"
        if choice == "5":
            return "html"
        if choice == "6":
            if not xlsx_available:
                warning("openpyxl not installed. Install it with: pip install openpyxl")
                continue
            return "xlsx"
        if choice == "7":
            return None

        warning("Invalid choice.")


def export_menu():
    return show_menu(
        "Export Center",
        [
            "AniList",
            "MAL",
            "Telegram",
            "Missing",
            "Retry Queue",
            "Aliases",
            "Search Cache",
            "Back",
        ],
    )


def settings_home():
    while True:
        choice = show_menu(
            "Settings",
            [
                "Basic Settings",
                "Advanced Settings",
                "Back",
            ],
        )

        if choice == "1":
            settings_editor("basic")
            continue

        if choice == "2":
            settings_editor("advanced")
            continue

        if choice == "3":
            return



def settings_editor(mode):
    from settings import load_settings

    settings = load_settings()

    while True:

        title = "Basic Settings" if mode == "basic" else "Advanced Settings"
        show_header(title)
        option_map = {}
        option = 1

        for section, items in SETTINGS_MENU:

            visible_items = []

            for key, label in items:

                if mode == "basic" and key not in BASIC_SETTINGS:
                    continue

                if mode == "advanced" and key not in ADVANCED_SETTINGS:
                    continue

                visible_items.append((key, label))

            if not visible_items:
                continue

            print(f"[ {section} ]")

            for key, label in visible_items:

                option_map[option] = key

                value = None

                if section == "Synchronization":
                    value = settings["sync"][key]
                elif section == "Search":
                    value = settings["search"][key]
                elif section == "Backup":
                    value = settings["backup"][key]
                elif section == "User Interface":
                    value = settings["ui"][key]
                elif section == "Advanced":
                    value = settings.get(key)

                if isinstance(value, bool):
                    value = "ON" if value else "OFF"

                print(f"{option}. {label:<30} {value}")

                option += 1

            print()

        back_option = option
        print("-" * 40)
        print(f"{back_option}. Back")
        print()

        pick = ask()
        if not pick.isdigit():
            warning("Invalid choice.")
            continue

        idx = int(pick)
        if idx == back_option:
            break

        key = option_map.get(idx)
        if not key:
            warning("Invalid choice.")
            continue

        # Determine which nested section the key belongs to and preserve type.
        if key in settings.get("sync", {}):
            section_key = "sync"
        elif key in settings.get("search", {}):
            section_key = "search"
        elif key in settings.get("backup", {}):
            section_key = "backup"
        elif key in settings.get("ui", {}):
            section_key = "ui"
        elif key in settings:
            section_key = None
        else:
            error("Unknown setting key.")
            continue

        if section_key is None:
            old_val = settings[key]
        else:
            old_val = settings[section_key][key]

        # Boolean settings toggle automatically
        if isinstance(old_val, bool):
            if section_key is None:
                settings[key] = not old_val
            else:
                settings[section_key][key] = not old_val

            # Display friendly label
            label = key
            for _, items in SETTINGS_MENU:
                for setting_key, setting_label in items:
                    if setting_key == key:
                        label = setting_label
                        break
                if label != key:
                    break

            save_json(SETTINGS_FILE, settings)

            success(f"{label} -> {'ON' if settings[section_key][key] else 'OFF'}")
            continue

        # Numeric/Text settings still ask for input
        new_val = ask(f"New value for {key}:")

        if isinstance(old_val, int):
            value = int(new_val)

        elif isinstance(old_val, float):
            value = float(new_val)

        else:
            value = new_val

        if section_key is None:
            settings[key] = value
        else:
            settings[section_key][key] = value

        save_json(SETTINGS_FILE, settings)
        success("Saved.")


def bulk_operations():
    while True:
        choice = show_menu(
            "Bulk Operations",
            [
                "Refresh all MAL IDs",
                "Refresh all AniList IDs",
                "Refresh Library Caches",
                "Refresh Search Cache",
                "Retry Failed Titles",
                "Remove Duplicate Aliases",
                "Clean Old Backups",
                "Rebuild Statistics",
                "Verify Library",
                "Back",
            ],
        )

        if choice == "1":
            from mal import get_completed_mal_ids
            print("Refreshing MAL IDs from API...")
            mal_ids = list(get_completed_mal_ids(force_refresh=True))
            with open("state.json", encoding="utf-8") as f:
                state = json.load(f)
            state["mal_ids"] = mal_ids
            with open("state.json", "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2)
            success(f"Refreshed {len(mal_ids)} MAL IDs.")

        elif choice == "2":
            from anilist import get_completed_ids
            print("Refreshing AniList IDs from API...")
            anilist_ids = list(get_completed_ids(force_refresh=True))
            with open("state.json", encoding="utf-8") as f:
                state = json.load(f)
            state["anilist_ids"] = anilist_ids
            with open("state.json", "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2)
            success(f"Refreshed {len(anilist_ids)} AniList IDs.")

        elif choice == "3":
            from anilist import get_completed_ids as get_anilist_ids
            from mal import get_completed_mal_ids
            print("Refreshing all library caches from API...")
            anilist_ids = list(get_anilist_ids(force_refresh=True))
            mal_ids = list(get_completed_mal_ids(force_refresh=True))
            with open("state.json", encoding="utf-8") as f:
                state = json.load(f)
            state["anilist_ids"] = anilist_ids
            state["mal_ids"] = mal_ids
            with open("state.json", "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2)
            success(f"Refreshed {len(anilist_ids)} AniList + {len(mal_ids)} MAL IDs.")

        elif choice == "4":
            from modes.search_cache import search_cache
            search_cache()

        elif choice == "5":
            from modes.retry_queue import retry_queue_menu
            retry_queue_menu()

        elif choice == "6":
            from modes.alias_manager import detect_duplicates
            detect_duplicates()

        elif choice == "7":
            _clean_old_backups()

        elif choice == "8":
            from modes.statistics import statistics
            statistics()

        elif choice == "9":
            library_health()

        elif choice == "10":
            break
        else:
            warning("Invalid choice.")


async def data_center():
    ensure_exports()

    while True:
        choice = show_menu(
            "Tools",
            [
                "Export",
                "Import",
                "Backup",
                "Restore",
                "Alias Manager",
                "Search Cache",
                "Retry Queue",
                "Settings",
                "Library Health",
                "Back",
            ],
        )

        if choice == "1":
            await export_center()
        elif choice == "2":
            import_center()
        elif choice == "3":
            backup_center()
        elif choice == "4":
            restore_center()
        elif choice == "5":
            from modes.alias_manager import alias_manager
            alias_manager()
        elif choice == "6":
            from modes.search_cache import search_cache_menu
            search_cache_menu()
        elif choice == "7":
            from modes.retry_queue import retry_queue_menu
            retry_queue_menu()

        elif choice == "8":
            settings_home()

        elif choice == "9":
            library_health()

        elif choice == "10":
            break
        else:
            warning("Invalid choice.")


async def export_center():
    while True:
        choice = export_menu()

        if choice == "1":
            export_anilist_library()
        elif choice == "2":
            export_mal_library()
        elif choice == "3":
            await export_telegram_titles()
        elif choice == "4":
            export_missing_anime()
        elif choice == "5":
            export_retry_queue()
        elif choice == "6":
            export_aliases()
        elif choice == "7":
            export_search_cache()
        elif choice == "8":
            break
        else:
            warning("Invalid choice.")


def export_anilist_library():
    from anilist import get_completed_anime

    anime = sorted(
        get_completed_anime(),
        key=lambda item: item.get("title", "").lower()
    )
    rows = _library_rows(anime)
    headers = ["Title", "AniList ID", "MAL ID", "Episodes", "Status", "Progress"]
    txt_lines = [item.get("title", "") for item in anime]

    _export_dataset("anilist_library", anime, rows, headers, txt_lines=txt_lines)


def export_mal_library():
    from mal import get_completed_mal_anime

    anime = sorted(
        get_completed_mal_anime(),
        key=lambda item: item.get("title", "").lower()
    )
    rows = _library_rows(anime)
    headers = ["Title", "AniList ID", "MAL ID", "Episodes", "Status", "Progress"]
    txt_lines = [item.get("title", "") for item in anime]

    _export_dataset("mal_library", anime, rows, headers, txt_lines=txt_lines)


async def export_telegram_titles():
    from telegram_client import client

    started_here = False
    seen = set()
    titles = []

    if not client.is_connected():
        await client.start()
        started_here = True

    try:
        async for message in client.iter_messages("me", reverse=True):
            if not getattr(message, "text", None):
                continue

            title = message.text.strip()
            if not title or title in seen:
                continue

            seen.add(title)
            titles.append(title)
    finally:
        if started_here:
            await client.disconnect()

    rows = [{"Title": title} for title in titles]
    _export_dataset("telegram_titles", titles, rows, ["Title"])


def export_missing_anime():
    report = load_json("missing_anilist.json", None)

    if report is None:
        root_report = Path("missing_anilist.json")
        if root_report.exists():
            with open(root_report, "r", encoding="utf-8") as f:
                report = json.load(f)

    if not report:
        warning("No missing anime report found. Run Compare first.")
        return

    missing = report.get("missing", [])
    rows = [
        {
            "Title": item.get("matched_title") or item.get("telegram_title", ""),
            "Telegram Title": item.get("telegram_title", ""),
            "AniList ID": item.get("id", ""),
            "MAL ID": item.get("idMal", ""),
            "Episodes": item.get("episodes", ""),
            "Reason": item.get("reason", ""),
        }
        for item in missing
    ]
    headers = ["Title", "Telegram Title", "AniList ID", "MAL ID", "Episodes", "Reason"]
    _export_dataset("missing", missing, rows, headers)


def export_retry_queue():
    queue = load_json(RETRY_FILE, [])
    rows = [{"Title": title} for title in queue]
    _export_dataset("retry_queue", queue, rows, ["Title"])


def export_aliases():
    aliases = load_json(ALIASES_FILE, {})
    rows = _alias_rows(aliases)
    headers = ["Alias", "Title", "AniList ID", "MAL ID", "Episodes"]

    _export_dataset("aliases", aliases, rows, headers, txt_lines=_alias_txt_lines(aliases))


def export_search_cache():
    cache = load_json(CACHE_FILE, {})
    rows = []

    for query, item in sorted(cache.items()):
        title, anilist_id, mal_id, episodes = _anime_fields(item)
        rows.append({
            "Query": query,
            "Title": title,
            "AniList ID": anilist_id,
            "MAL ID": mal_id,
            "Episodes": episodes,
        })

    headers = ["Query", "Title", "AniList ID", "MAL ID", "Episodes"]
    _export_dataset("search_cache", cache, rows, headers)


def import_center():
    ensure_exports()

    while True:
        choice = show_menu(
            "Import Center",
            [
                "aliases.json",
                "retry_queue.json",
                "search_cache.json",
                "settings.json",
                "telegram.txt",
                "Custom file",
                "Back",
            ],
        )

        if choice == "1":
            import_json_file(ALIASES_FILE, merge=True)
        elif choice == "2":
            import_json_file(RETRY_FILE, merge=True)
        elif choice == "3":
            import_json_file(CACHE_FILE, merge=True)
        elif choice == "4":
            import_json_file(SETTINGS_FILE, merge=False)
        elif choice == "5":
            import_telegram_txt("telegram.txt")
        elif choice == "6":
            import_custom_file()
        elif choice == "7":
            break
        else:
            warning("Invalid choice.")


def backup_center():
    created = 0

    for filename in DATA_FILES:
        if data_file(filename).exists():
            backup_file(filename)
            created += 1

    success(f"Backed up {created} data file(s).")


def restore_center():
    backups = sorted(Path(BACKUP_DIR).glob("*.json"), reverse=True)

    if not backups:
        warning("No backups found.")
        return

    choice = show_menu(
        "Restore Backup",
        [path.name for path in backups] + ["Back"],
    )

    if not choice.isdigit():
        warning("Invalid choice.")
        return

    index = int(choice)
    if index == len(backups) + 1:
        return
    if index < 1 or index > len(backups):
        warning("Invalid choice.")
        return

    backup = backups[index - 1]
    original = _original_backup_name(backup.name)

    if not original:
        error("Could not detect original filename.")
        return

    confirm = ask(f"Restore {backup.name} to data/{original}? (y/n):").lower()
    if confirm != "y":
        warning("Restore cancelled.")
        return

    if data_file(original).exists():
        backup_file(original)

    copy2(backup, data_file(original))
    success(f"Restored data/{original}.")


def import_json_file(filename, merge=False, path=None):
    if path is None:
        path = _ask_import_path(filename)

    if not path:
        return

    try:
        with open(path, "r", encoding="utf-8") as f:
            incoming = json.load(f)
    except Exception as exc:
        error(f"Could not read JSON: {exc}")
        return

    current = load_json(filename, [] if isinstance(incoming, list) else {})

    if merge:
        mode = ask("Merge with current data? (Y=merge, N=replace, C=cancel):").lower()
        if mode == "c":
            warning("Import cancelled.")
            return
        replace = mode == "n"
    else:
        confirm = ask(f"Replace data/{filename}? (y/n):").lower()
        if confirm != "y":
            warning("Import cancelled.")
            return
        replace = True

    backup_file(filename)

    if replace:
        save_json(filename, incoming)
        success(f"Imported {filename}.")
        return

    merged = _merge_data(current, incoming)
    save_json(filename, merged)
    success(f"Merged import into {filename}.")


def import_telegram_txt(filename, path=None):
    if path is None:
        path = _ask_import_path(filename)

    if not path:
        return

    with open(path, "r", encoding="utf-8") as f:
        titles = [line.strip() for line in f if line.strip()]

    if not titles:
        warning("No titles found.")
        return

    queue = load_json(RETRY_FILE, [])
    added = 0

    for title in titles:
        if title in queue:
            continue
        queue.append(title)
        added += 1

    backup_file(RETRY_FILE)
    save_json(RETRY_FILE, queue)
    success(f"Imported {added} title(s) into retry_queue.json.")


def import_custom_file():
    raw = ask("File path:").strip('"')
    if not raw:
        return

    path = Path(raw)
    if not path.exists():
        path = EXPORT_PATH / raw

    if not path.exists():
        error("File not found.")
        return

    name = path.name.lower()

    if name == ALIASES_FILE:
        import_json_file(ALIASES_FILE, merge=True, path=path)
    elif name == RETRY_FILE:
        import_json_file(RETRY_FILE, merge=True, path=path)
    elif name == CACHE_FILE:
        import_json_file(CACHE_FILE, merge=True, path=path)
    elif name == SETTINGS_FILE:
        import_json_file(SETTINGS_FILE, merge=False, path=path)
    elif name.endswith(".txt"):
        import_telegram_txt(path.name, path=path)
    else:
        warning("Supported custom imports: aliases, retry queue, cache, settings, or TXT titles.")


def _export_dataset(name, json_data, rows, headers, txt_lines=None):
    fmt = choose_format()

    if not fmt:
        warning("Export cancelled.")
        return

    if fmt == "json":
        path = export_json(name, json_data)
    elif fmt == "csv":
        path = export_csv(name, rows, headers)
    elif fmt == "txt":
        path = export_txt(name, txt_lines if txt_lines is not None else rows)
    elif fmt == "html":
        path = export_html(name, rows, headers)
    elif fmt == "xlsx":
        path = export_xlsx(name, rows, headers)
        if not path:
            return
    else:
        path = export_markdown(name, rows, headers)

    success(f"Exported to {path}")


def _alias_rows(aliases):
    rows = []

    for alias, data in sorted(aliases.items()):
        rows.append({
            "Alias": alias,
            "Title": data.get("title", ""),
            "AniList ID": data.get("id", ""),
            "MAL ID": data.get("idMal", ""),
            "Episodes": data.get("episodes", ""),
        })

    return rows


def _library_rows(anime_list):
    rows = []

    for anime in anime_list:
        rows.append({
            "Title": anime.get("title", ""),
            "AniList ID": anime.get("id", ""),
            "MAL ID": anime.get("idMal", ""),
            "Episodes": anime.get("episodes", ""),
            "Status": anime.get("status", ""),
            "Progress": anime.get("progress", ""),
        })

    return rows


def _alias_txt_lines(aliases):
    lines = []

    for alias, data in sorted(aliases.items()):
        lines.append(
            f"{alias}\n->\n{data.get('title', '')}\n----------------"
        )

    return lines


def _anime_fields(item):
    if not item:
        return "", "", "", ""

    if isinstance(item, list):
        item = item[0][1] if item and isinstance(item[0], list) and len(item[0]) > 1 else item[0]

    title = ""
    if isinstance(item, dict):
        raw_title = item.get("title", "")
        if isinstance(raw_title, dict):
            title = (
                raw_title.get("english")
                or raw_title.get("romaji")
                or raw_title.get("native")
                or ""
            )
        else:
            title = raw_title

        return (
            title,
            item.get("id", ""),
            item.get("idMal", ""),
            item.get("episodes", ""),
        )

    return str(item), "", "", ""


def _txt_lines(data):
    if isinstance(data, list):
        lines = []
        for item in data:
            if isinstance(item, dict):
                lines.append(item.get("Title") or item.get("Alias") or str(item))
            else:
                lines.append(str(item))
        return lines

    if isinstance(data, dict):
        return [f"{key}: {value}" for key, value in data.items()]

    return [str(data)]


def _merge_data(current, incoming):
    if isinstance(current, dict) and isinstance(incoming, dict):
        merged = dict(current)
        merged.update(incoming)
        return merged

    if isinstance(current, list) and isinstance(incoming, list):
        merged = list(current)
        for item in incoming:
            if item not in merged:
                merged.append(item)
        return merged

    return incoming


def _ask_import_path(default_name):
    raw = ask(f"Import file [{default_name}]:").strip('"')
    path = Path(raw or default_name)

    if not path.exists():
        path = EXPORT_PATH / path

    if not path.exists():
        error("File not found.")
        return None

    return path


def _original_backup_name(filename):
    match = re.match(r"(.+)_\d{8}_\d{6}(\.json)$", filename)
    if not match:
        return None

    return match.group(1) + match.group(2)


def _with_suffix(filename, suffix):
    path = Path(filename)

    if path.suffix:
        return path.name

    return f"{filename}.{suffix}"


def _clean_old_backups(keep: int = 50):
    """Remove oldest backups, keeping the most recent `keep`."""

    path = Path(BACKUP_DIR)
    if not path.is_dir():
        warning("No backup directory found.")
        return

    backups = sorted(path.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
    if len(backups) <= keep:
        success(f"Only {len(backups)} backups, nothing to clean.")
        return

    to_remove = backups[keep:]
    for b in to_remove:
        try:
            b.unlink()
        except Exception:
            pass

    success(f"Removed {len(to_remove)} old backups, kept {keep}.")


def _health_input():
    """Read input with ESC support. Returns the string or None if ESC pressed."""
    buf = ""
    while True:
        if msvcrt.kbhit():
            key = msvcrt.getch()
            if key == b"\x1b":
                return None
            if key == b"\r":
                return buf
            if key == b"\x7f" or key == b"\x08":
                buf = buf[:-1]
                print("\b \b", end="", flush=True)
            elif key in (b"\xe0", b"\x00"):
                msvcrt.getch()
            else:
                try:
                    ch = key.decode("utf-8")
                    buf += ch
                    print(ch, end="", flush=True)
                except UnicodeDecodeError:
                    pass

def _export_health_report(pct, groups, issues):
    from datetime import datetime

    name = f"health_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    rows = []
    for group_name, items in groups:
        for name, status in items:
            rows.append({"Section": f"{group_name} / {name}", "Status": status})
    if issues:
        for issue in issues:
            rows.append({"Section": "Suggestion", "Status": issue})
    rows.append({"Section": "Overall", "Status": f"{pct}%"})

    json_data = {
        "health_pct": pct,
        "timestamp": datetime.now().isoformat(),
        "groups": [
            {
                "group": group_name,
                "checks": [{"name": n, "status": s} for n, s in items],
            }
            for group_name, items in groups
        ],
        "issues": issues,
    }

    _export_dataset(name, json_data, rows, ["Section", "Status"])


def _compute_health_score() -> tuple:
    """Compute health score without any output. Returns (pct, groups, issues)."""
    from settings import DEFAULT_SETTINGS

    issues = []
    total_checks = 12
    passed = 0

    aliases = load_json(ALIASES_FILE, {})
    broken = [k for k, v in aliases.items() if not v or not v.get("id")]
    dup_count = 0
    seen = set()
    for k, v in aliases.items():
        key = re.sub(r"[^a-zA-Z0-9]", "", k).lower()
        if key in seen:
            dup_count += 1
        seen.add(key)
    if not broken and not dup_count:
        passed += 1
    if not dup_count:
        passed += 1

    cache = load_json(CACHE_FILE, {})
    cache_age_days = None
    try:
        mtime = Path(CACHE_FILE).stat().st_mtime
        cache_age_days = int((time.time() - mtime) / 86400)
    except Exception:
        pass
    if cache and (cache_age_days is None or cache_age_days <= 30):
        passed += 1

    retry = load_json(RETRY_FILE, [])
    if not retry:
        passed += 1

    resume = load_json(RESUME_FILE, {})
    resume_ok = True
    if not resume:
        resume_ok = False
    else:
        msg_id = resume.get("last_message_id")
        if msg_id is None or not isinstance(msg_id, int) or msg_id < 0:
            resume_ok = False
    if resume_ok:
        passed += 1

    backup_count = 0
    try:
        backup_path = Path(BACKUP_DIR)
        if backup_path.is_dir():
            backup_count = len(list(backup_path.iterdir()))
    except Exception:
        pass
    if backup_count <= 100:
        passed += 1

    export_issues = 0
    export_path = Path(EXPORT_DIR)
    if not export_path.is_dir():
        export_issues += 1
    else:
        export_files = list(export_path.iterdir())
        if not export_files:
            export_issues += 1
        else:
            corrupted = 0
            for f in export_files:
                if f.suffix == ".json":
                    try:
                        json.loads(f.read_text(encoding="utf-8"))
                    except Exception:
                        corrupted += 1
            if corrupted:
                export_issues += 1
    if not export_issues:
        passed += 1

    settings = load_json(SETTINGS_FILE, {})
    config_issues = 0
    if not isinstance(settings, dict):
        config_issues += 1
    else:
        missing = [k for k in DEFAULT_SETTINGS if k not in settings]
        if missing:
            config_issues += 1
        unknown = [k for k in settings if k not in DEFAULT_SETTINGS]
        if unknown:
            config_issues += 1
        invalid = []
        for k, v in DEFAULT_SETTINGS.items():
            if k in settings and settings[k] is not None:
                if not isinstance(settings[k], type(v)):
                    invalid.append(k)
        if invalid:
            config_issues += 1
    if not config_issues:
        passed += 1

    missing_mal = 0
    try:
        with open("state.json", encoding="utf-8") as f:
            state = json.load(f)
        mal_ids = state.get("mal_ids", [])
        missing_mal = sum(1 for mid in mal_ids if not mid)
        if not missing_mal:
            passed += 1
    except Exception:
        pass

    from config import ANILIST_TOKEN
    cred_issues = 0
    if not ANILIST_TOKEN or ANILIST_TOKEN == "your_anilist_access_token":
        cred_issues += 1
    else:
        try:
            r = requests.post(
                "https://graphql.anilist.co",
                json={"query": "{ Viewer { id } }"},
                headers={"Authorization": f"Bearer {ANILIST_TOKEN}"},
                timeout=10,
            )
            if r.status_code != 200:
                cred_issues += 1
        except Exception:
            cred_issues += 1
    mal_tokens = load_json("mal_tokens.json", {})
    if not mal_tokens or not mal_tokens.get("access_token"):
        cred_issues += 1
    else:
        expires = mal_tokens.get("expires_at", 0)
        if time.time() >= expires:
            cred_issues += 1
    if not cred_issues:
        passed += 1

    from config import API_ID, API_HASH
    telegram_ok = True
    if not API_ID or API_ID == 0:
        telegram_ok = False
    if not API_HASH or API_HASH == "your_telegram_api_hash":
        telegram_ok = False
    if not Path("telegram_session.session").exists():
        telegram_ok = False
    if telegram_ok:
        passed += 1

    anilist_ok = True
    try:
        with open("state.json", encoding="utf-8") as f:
            state = json.load(f)
        anilist_ids = state.get("anilist_ids", [])
        if not anilist_ids:
            anilist_ok = False
    except Exception:
        anilist_ok = False
    if anilist_ok:
        passed += 1

    pct = int(passed / total_checks * 100) if total_checks else 100

    groups = [
        ("Library", [
            ("Aliases",
             "✓ OK" if not broken and not dup_count
             else f"⚠ {len(broken)} broken, {dup_count} dup" if broken and dup_count
             else f"⚠ {len(broken)} broken" if broken
             else f"⚠ {dup_count} dup"),
            ("Retry Queue", "✓ Empty" if not retry else f"⚠ {len(retry)} pending"),
        ]),
        ("Storage", [
            ("Search Cache",
             f"✓ {len(cache)} entries" if cache and (cache_age_days is None or cache_age_days <= 30)
             else f"⚠ {len(cache)} entries, {cache_age_days}d" if cache
             else "⚠ Empty"),
            ("Exports", "✓ OK" if not export_issues else "⚠ Issues"),
            ("Backups",
             f"✓ {backup_count}" if backup_count <= 100 else f"⚠ {backup_count}"),
        ]),
        ("Accounts", [
            ("API Credentials", "✓ OK" if not cred_issues else "⚠ Issues"),
            ("Telegram", "✓ OK" if telegram_ok else "⚠ Issues"),
            ("AniList", "✓ OK" if anilist_ok else "⚠ Missing"),
            ("MyAnimeList", "✓ OK" if not missing_mal else f"⚠ {missing_mal} missing"),
        ]),
        ("Configuration", [
            ("Settings", "✓ OK" if not config_issues else "⚠ Issues"),
            ("Resume File", "✓ OK" if resume_ok else "⚠ Issues"),
        ]),
    ]

    # Rebuild the issues list for the menu
    issues = []
    if broken:
        issues.append(f"⚠ {len(broken)} broken aliases")
    if dup_count:
        issues.append(f"⚠ {dup_count} duplicate aliases")
    if cache and cache_age_days is not None and cache_age_days > 30:
        issues.append(f"⚠ Cache hasn't been refreshed in {cache_age_days} days ({len(cache)} entries)")
    elif not cache:
        issues.append("⚠ Search cache is empty")
    if retry:
        issues.append(f"⚠ {len(retry)} entries in retry queue")
    if not resume:
        issues.append("⚠ Resume file is missing or empty")
    elif not resume_ok:
        issues.append("⚠ Resume file has invalid last_message_id")
    if backup_count > 100:
        issues.append(f"⚠ Large backup folder ({backup_count} backups)")
    if export_issues:
        if not export_path.is_dir():
            issues.append("⚠ Export folder is missing")
        elif not list(export_path.iterdir()):
            issues.append("⚠ Export folder is empty")
        else:
            issues.append(f"⚠ {corrupted} corrupted export files")
    if not isinstance(settings, dict):
        issues.append("⚠ Settings file is corrupted")
    else:
        if missing:
            issues.append(f"⚠ {len(missing)} missing settings")
        if unknown:
            issues.append(f"⚠ {len(unknown)} unknown settings keys")
        if invalid:
            issues.append(f"⚠ {len(invalid)} settings with wrong type")
    if missing_mal:
        issues.append(f"⚠ {missing_mal} entries missing MAL IDs")
    if cred_issues:
        if not ANILIST_TOKEN or ANILIST_TOKEN == "your_anilist_access_token":
            issues.append("⚠ AniList token missing or placeholder")
        else:
            issues.append("⚠ AniList token is invalid or expired")
        if not mal_tokens or not mal_tokens.get("access_token"):
            issues.append("⚠ MAL tokens missing")
        elif time.time() >= mal_tokens.get("expires_at", 0):
            issues.append("⚠ MAL token is expired")
    if not telegram_ok:
        if not API_ID or API_ID == 0:
            issues.append("⚠ Telegram API_ID not configured")
        if not API_HASH or API_HASH == "your_telegram_api_hash":
            issues.append("⚠ Telegram API_HASH not configured")
        if not Path("telegram_session.session").exists():
            issues.append("⚠ Telegram session file missing")
    if not anilist_ok:
        issues.append("⚠ AniList library not loaded")

    return pct, groups, issues


def library_health():
    show_header("Library Health")
    print()

    pct, groups, issues = _compute_health_score()
    settings = load_json(SETTINGS_FILE, {})
    color = "🟢" if pct >= 80 else ("🟡" if pct >= 50 else "🔴")

    show_header(f"Library Health — {pct}% {color}")
    print()

    for group_name, items in groups:
        print(group_name)
        print("─" * 40)
        for name, status in items:
            print(f"  {status}  {name}")
        print()

    while True:
        if issues:
            print("Suggestions")
            print("─" * 40)
            for i, issue in enumerate(issues, 1):
                if "duplicate" in issue.lower():
                    label = "Merge duplicate aliases"
                elif "mal id" in issue.lower():
                    label = "Repair missing MAL IDs"
                elif "retry" in issue.lower():
                    label = "Retry failed titles"
                elif "broken" in issue.lower():
                    label = "Fix broken aliases"
                elif "cache" in issue.lower():
                    label = "Refresh stale cache"
                elif "backup" in issue.lower():
                    label = "Clean old backups"
                elif "resume" in issue.lower():
                    label = "Review resume file"
                elif "export" in issue.lower() or "corrupted" in issue.lower():
                    label = "Review exports"
                elif "missing setting" in issue.lower():
                    label = "Add missing settings"
                elif "unknown setting" in issue.lower():
                    label = "Review unknown settings"
                elif "wrong type" in issue.lower():
                    label = "Fix setting types"
                elif "token" in issue.lower():
                    label = "Fix API credentials"
                elif "telegram" in issue.lower():
                    label = "Fix Telegram connection"
                elif "anilist library" in issue.lower():
                    label = "Sync AniList library"
                else:
                    label = issue
                print(f"  {i}. {label}")
            print()

        print("  0. Export report")
        warning("Press ESC to return.")
        print()
        print("Fix: ", end="", flush=True)

        choice = _health_input()
        print()
        if choice is None:
            return
        if choice == "0":
            _export_health_report(pct, groups, issues)
            continue
        if choice.isdigit() and issues:
            idx = int(choice) - 1
            issue_types = []
            for issue in issues:
                if "duplicate" in issue.lower():
                    issue_types.append("aliases")
                elif "mal id" in issue.lower():
                    issue_types.append("repair")
                elif "retry" in issue.lower():
                    issue_types.append("retry")
                elif "broken" in issue.lower():
                    issue_types.append("aliases")
                elif "cache" in issue.lower():
                    issue_types.append("cache")
                elif "backup" in issue.lower():
                    issue_types.append("backup")
                elif "resume" in issue.lower():
                    issue_types.append("resume")
                elif "export" in issue.lower() or "corrupted" in issue.lower():
                    issue_types.append("export")
                elif "missing setting" in issue.lower() or "wrong type" in issue.lower():
                    issue_types.append("config")
                elif "token" in issue.lower():
                    issue_types.append("creds")
                elif "telegram" in issue.lower():
                    issue_types.append("telegram")
                elif "anilist library" in issue.lower():
                    issue_types.append("sync")
                else:
                    issue_types.append(None)
            if 0 <= idx < len(issue_types):
                action = issue_types[idx]
                if action == "aliases":
                    from modes.alias_manager import detect_duplicates
                    detect_duplicates()
                    library_health()
                    return
                elif action == "repair":
                    from modes.repair import repair as run_repair
                    run_repair()
                    library_health()
                    return
                elif action == "retry":
                    from modes.retry_queue import retry_queue_menu
                    retry_queue_menu()
                    library_health()
                    return
                elif action == "cache":
                    from modes.search_cache import search_cache_menu
                    search_cache_menu()
                    library_health()
                    return
                elif action == "backup":
                    _clean_old_backups()
                    library_health()
                    return
                elif action == "resume":
                    save_json(RESUME_FILE, {"last_message_id": 0})
                    success("Resume file reset to last_message_id: 0.")
                    library_health()
                    return
                elif action == "export":
                    warning("No automated fix — review exports/ folder manually.")
                    continue
                elif action == "config":
                    from settings import DEFAULT_SETTINGS
                    fixed = dict(settings)
                    for k, v in DEFAULT_SETTINGS.items():
                        fixed.setdefault(k, v)
                    fixed = {k: v for k, v in fixed.items() if k in DEFAULT_SETTINGS}
                    for k, v in DEFAULT_SETTINGS.items():
                        if k in fixed and fixed[k] is not None and not isinstance(fixed[k], type(v)):
                            fixed[k] = v
                    save_json(SETTINGS_FILE, fixed)
                    success("Configuration repaired.")
                    library_health()
                    return
                elif action == "creds":
                    warning("No automated fix — check config.py and mal_tokens.json manually.")
                    continue
                elif action == "telegram":
                    warning("No automated fix — check config.py and telegram_session.session manually.")
                    continue
                elif action == "sync":
                    warning("Run a full sync to populate library data.")
                    continue




