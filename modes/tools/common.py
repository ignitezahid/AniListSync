import csv
import json
from pathlib import Path

from utils.constants import EXPORT_DIR
from utils.ui import success, warning, show_menu
from utils.menu_keys import *  # noqa: F405


EXPORT_PATH = Path(EXPORT_DIR)

DATA_FILES = [
    "aliases.json",
    "search_cache.json",
    "retry_queue.json",
    "resume.json",
    "settings.json",
    "missing_anilist.json",
]


def ensure_exports():
    EXPORT_PATH.mkdir(parents=True, exist_ok=True)


def export_path(filename):
    ensure_exports()
    return EXPORT_PATH / filename


def _with_suffix(filename, suffix):
    path = Path(filename)
    if path.suffix:
        return path.name
    return f"{filename}.{suffix}"


def export_json(filename, data):
    path = export_path(_with_suffix(filename, "json"))
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
    return path


def export_txt(filename, data):
    path = export_path(_with_suffix(filename, "txt"))
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(_txt_lines(data)))
        f.write("\n")
    return path


def export_csv(filename, rows, headers):
    path = export_path(_with_suffix(filename, "csv"))
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return path


def export_markdown(filename, rows, headers):
    path = export_path(_with_suffix(filename, "md"))
    with open(path, "w", encoding="utf-8") as f:
        f.write("| " + " | ".join(headers) + " |\n")
        f.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
        for row in rows:
            values = [str(row.get(header, "")).replace("\n", " ") for header in headers]
            f.write("| " + " | ".join(values) + " |\n")
    return path


def export_html(filename, rows, headers):
    path = export_path(_with_suffix(filename, "html"))
    with open(path, "w", encoding="utf-8") as f:
        f.write("<!DOCTYPE html>\n<html>\n<head>\n")
        f.write('<meta charset="utf-8">\n')
        f.write("<style>\n")
        f.write("body { font-family: -apple-system, sans-serif; margin: 20px; }\n")
        f.write("table { border-collapse: collapse; width: 100%; }\n")
        f.write("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }\n")
        f.write("th { background: #4a90d9; color: white; }\n")
        f.write("tr:nth-child(even) { background: #f5f5f5; }\n")
        f.write("</style>\n</head>\n<body>\n")
        f.write(f"<h2>{filename}</h2>\n")
        f.write("<table>\n<thead>\n<tr>")
        for h in headers:
            f.write(f"<th>{h}</th>")
        f.write("</tr>\n</thead>\n<tbody>\n")
        for row in rows:
            f.write("<tr>")
            for h in headers:
                val = str(row.get(h, "")).replace("\n", "<br>")
                f.write(f"<td>{val}</td>")
            f.write("</tr>\n")
        f.write("</tbody>\n</table>\n</body>\n</html>\n")
    return path


def export_xlsx(filename, rows, headers):
    path = export_path(_with_suffix(filename, "xlsx"))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        warning("openpyxl not installed. Install it with: pip install openpyxl")
        return None
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 4)
    wb.save(path)
    return path


def choose_format():
    try:
        import openpyxl  # noqa: F401
        xlsx_available = True
    except ImportError:
        xlsx_available = False

    while True:
        options = [
            "JSON",
            "CSV",
            "TXT",
            "Markdown",
            "HTML",
            "Excel (.xlsx)" + ("" if xlsx_available else " [dim](requires: pip install openpyxl)[/]"),
            "Cancel",
        ]
        choice = show_menu("Export Format", options)
        if choice == FMT_JSON:
            return "json"
        if choice == FMT_CSV:
            return "csv"
        if choice == FMT_TXT:
            return "txt"
        if choice == FMT_MD:
            return "md"
        if choice == FMT_HTML:
            return "html"
        if choice == FMT_XLSX:
            if not xlsx_available:
                warning("openpyxl not installed. Install it with: pip install openpyxl")
                continue
            return "xlsx"
        if choice == FMT_CANCEL:
            return None
        warning("Invalid choice.")


def _library_rows(anime_list):
    rows = []
    for anime in anime_list:
        rows.append({
            "Title": anime.get("title", ""),
            "AniList ID": anime.get("id", ""),
            "MAL ID": anime.get("idMal", ""),
            "Episodes": anime.get("episodes", ""),
            "Status": anime.get("status", ""),
            "Progress": anime.get("progress", ""),
        })
    return rows


def _alias_rows(aliases):
    rows = []
    for alias, data in sorted(aliases.items()):
        rows.append({
            "Alias": alias,
            "Title": data.get("title", ""),
            "AniList ID": data.get("id", ""),
            "MAL ID": data.get("idMal", ""),
            "Episodes": data.get("episodes", ""),
        })
    return rows


def _alias_txt_lines(aliases):
    lines = []
    for alias, data in sorted(aliases.items()):
        lines.append(f"{alias}\n->\n{data.get('title', '')}\n----------------")
    return lines


def _anime_fields(item):
    if not item:
        return "", "", "", ""
    if isinstance(item, list):
        item = item[0][1] if item and isinstance(item[0], list) and len(item[0]) > 1 else item[0]
    title = ""
    if isinstance(item, dict):
        raw_title = item.get("title", "")
        if isinstance(raw_title, dict):
            title = raw_title.get("english") or raw_title.get("romaji") or raw_title.get("native") or ""
        else:
            title = raw_title
        return title, item.get("id", ""), item.get("idMal", ""), item.get("episodes", "")
    return str(item), "", "", ""


def _txt_lines(data):
    if isinstance(data, list):
        lines = []
        for item in data:
            if isinstance(item, dict):
                lines.append(item.get("Title") or item.get("Alias") or str(item))
            else:
                lines.append(str(item))
        return lines
    if isinstance(data, dict):
        return [f"{key}: {value}" for key, value in data.items()]
    return [str(data)]


def _merge_data(current, incoming):
    if isinstance(current, dict) and isinstance(incoming, dict):
        merged = dict(current)
        merged.update(incoming)
        return merged
    if isinstance(current, list) and isinstance(incoming, list):
        merged = list(current)
        for item in incoming:
            if item not in merged:
                merged.append(item)
        return merged
    return incoming


def _export_dataset(name, json_data, rows, headers, txt_lines=None):
    fmt = choose_format()
    if not fmt:
        warning("Export cancelled.")
        return
    if fmt == "json":
        path = export_json(name, json_data)
    elif fmt == "csv":
        path = export_csv(name, rows, headers)
    elif fmt == "txt":
        path = export_txt(name, txt_lines if txt_lines is not None else rows)
    elif fmt == "html":
        path = export_html(name, rows, headers)
    elif fmt == "xlsx":
        path = export_xlsx(name, rows, headers)
        if not path:
            return
    else:
        path = export_markdown(name, rows, headers)
    success(f"Exported to {path}")
